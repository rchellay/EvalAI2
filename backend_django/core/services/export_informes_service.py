"""
Servicios de exportación PDF y Excel para Informes Inteligentes 2.0
"""

from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


class PDFExportService:
    """
    Genera PDFs formales para informes de grupo e individuales.
    """
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configurar estilos personalizados para el PDF."""
        
        # Título principal
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e3a8a'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Subtítulo
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        ))
        
        # Sección
        self.styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#6366f1'),
            spaceAfter=8,
            fontName='Helvetica-Bold'
        ))
        
        # Texto normal
        self.styles.add(ParagraphStyle(
            name='CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=14,
            alignment=TA_JUSTIFY,
            fontName='Helvetica'
        ))
        
        # Comentario
        self.styles.add(ParagraphStyle(
            name='Comment',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=14,
            alignment=TA_JUSTIFY,
            fontName='Helvetica',
            leftIndent=20,
            rightIndent=20,
            spaceBefore=6,
            spaceAfter=6
        ))
    
    def generar_pdf_grupo(self, data, trimestre, grupo_nombre):
        """
        Genera un PDF del informe de grupo.
        
        Args:
            data: Diccionario con los datos del informe de grupo
            trimestre: T1, T2 o T3
            grupo_nombre: Nombre del grupo
            
        Returns:
            BytesIO con el PDF generado
        """
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        story = []
        
        # Encabezado
        story.append(Paragraph(
            f"Informe General de Grupo - {grupo_nombre}",
            self.styles['CustomTitle']
        ))
        story.append(Paragraph(
            f"Trimestre: {trimestre} • {datetime.now().strftime('%d/%m/%Y')}",
            self.styles['CustomSubtitle']
        ))
        story.append(Spacer(1, 0.5*cm))
        
        # Resumen General
        story.append(Paragraph("📊 Rendimiento del Grupo", self.styles['SectionTitle']))
        
        resumen_data = [
            ['Indicador', 'Valor'],
            ['Total de Estudiantes', str(data.get('total_estudiantes', 0))],
            ['Media Global', f"{data.get('media_global', 0):.2f}"],
            ['Tasa de Aprobados', f"{data.get('tasa_aprobados', 0):.1f}%"],
            ['Asistencia Media', f"{data.get('asistencia_media', 0):.1f}%"],
        ]
        
        resumen_table = Table(resumen_data, colWidths=[8*cm, 8*cm])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(resumen_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Distribución de Notas
        story.append(Paragraph("📈 Distribución de Notas", self.styles['SectionTitle']))
        
        distribucion_data = [['Categoría', 'Cantidad', 'Porcentaje']]
        for dist in data.get('distribucion_notas', []):
            distribucion_data.append([
                dist['categoria'],
                str(dist['cantidad']),
                f"{dist['porcentaje']}%"
            ])
        
        distribucion_table = Table(distribucion_data, colWidths=[6*cm, 5*cm, 5*cm])
        distribucion_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(distribucion_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Medias por Asignatura
        story.append(Paragraph("📚 Media por Asignatura", self.styles['SectionTitle']))
        
        asignaturas_data = [['Asignatura', 'Media', 'Tendencia']]
        for asig in data.get('medias_por_asignatura', []):
            tendencia = "↑" if asig['tendencia'] > 0 else "↓" if asig['tendencia'] < 0 else "→"
            asignaturas_data.append([
                asig['nombre'],
                f"{asig['media']:.2f}",
                f"{tendencia} {abs(asig['tendencia']):.2f}"
            ])
        
        asignaturas_table = Table(asignaturas_data, colWidths=[8*cm, 4*cm, 4*cm])
        asignaturas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(asignaturas_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Áreas Destacadas
        story.append(Paragraph("🚀 Áreas Destacadas", self.styles['SectionTitle']))
        for area in data.get('areas_destacadas', []):
            story.append(Paragraph(f"• {area}", self.styles['CustomNormal']))
        story.append(Spacer(1, 0.3*cm))
        
        # Áreas de Mejora
        story.append(Paragraph("🚨 Áreas de Mejora", self.styles['SectionTitle']))
        for area in data.get('areas_mejora', []):
            story.append(Paragraph(f"• {area}", self.styles['CustomNormal']))
        story.append(Spacer(1, 0.5*cm))
        
        # Asistencia
        if data.get('ranking_absentismo'):
            story.append(Paragraph("🚦 Estudiantes con Mayor Absentismo", self.styles['SectionTitle']))
            absentismo_data = [['Posición', 'Estudiante', 'Horas de Ausencia']]
            for idx, est in enumerate(data['ranking_absentismo'][:5], 1):
                absentismo_data.append([
                    str(idx),
                    est['nombre'],
                    f"{est['horas_falta']}h"
                ])
            
            absentismo_table = Table(absentismo_data, colWidths=[3*cm, 9*cm, 4*cm])
            absentismo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(absentismo_table)
        
        # Footer
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph(
            f"Informe generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
            self.styles['CustomNormal']
        ))
        story.append(Paragraph(
            "EvalAI - Informes Inteligentes 2.0",
            self.styles['CustomNormal']
        ))
        
        # Generar PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def generar_pdf_individual(self, data, comentarios, trimestre, estudiante_nombre, grupo_nombre):
        """
        Genera un PDF del informe individual con comentarios de IA.
        
        Args:
            data: Diccionario con los datos del estudiante
            comentarios: Diccionario con los comentarios generados por IA
            trimestre: T1, T2 o T3
            estudiante_nombre: Nombre del estudiante
            grupo_nombre: Nombre del grupo
            
        Returns:
            BytesIO con el PDF generado
        """
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        story = []
        
        # Encabezado
        story.append(Paragraph(
            f"Informe Individual: {estudiante_nombre}",
            self.styles['CustomTitle']
        ))
        story.append(Paragraph(
            f"{grupo_nombre} • {trimestre} • {datetime.now().strftime('%d/%m/%Y')}",
            self.styles['CustomSubtitle']
        ))
        story.append(Spacer(1, 0.5*cm))
        
        # Datos del Alumno
        story.append(Paragraph("🔹 1. Datos del Alumno", self.styles['SectionTitle']))
        datos_table = Table([
            ['Nombre:', estudiante_nombre],
            ['Grupo:', grupo_nombre],
            ['Trimestre:', trimestre]
        ], colWidths=[4*cm, 12*cm])
        datos_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        story.append(datos_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Ausencias
        story.append(Paragraph("🔹 2. Ausencias", self.styles['SectionTitle']))
        story.append(Paragraph(
            f"Total de horas de ausencia: {data.get('total_horas_ausencia', 0)}h",
            self.styles['CustomNormal']
        ))
        story.append(Spacer(1, 0.5*cm))
        
        # Comentario General del Tutor
        story.append(Paragraph("🔹 3. Comentario General del Tutor", self.styles['SectionTitle']))
        if comentarios.get('comentario_general'):
            story.append(Paragraph(
                comentarios['comentario_general'],
                self.styles['Comment']
            ))
        else:
            story.append(Paragraph(
                "Comentario general pendiente.",
                self.styles['Comment']
            ))
        story.append(Spacer(1, 0.5*cm))
        
        # Autoevaluación
        if data.get('autoevaluacion'):
            story.append(Paragraph("🔹 4. Autoevaluación del Alumno", self.styles['SectionTitle']))
            story.append(Paragraph(
                f"\"{data['autoevaluacion']['texto']}\"",
                self.styles['Comment']
            ))
            
            if comentarios.get('comentario_autoevaluacion'):
                story.append(Paragraph(
                    "<b>Comentario del tutor sobre la autoevaluación:</b>",
                    self.styles['CustomNormal']
                ))
                story.append(Paragraph(
                    comentarios['comentario_autoevaluacion'],
                    self.styles['Comment']
                ))
            story.append(Spacer(1, 0.5*cm))
        
        # Rendimiento por Asignatura
        story.append(Paragraph("🔹 5. Rendimiento por Asignatura", self.styles['SectionTitle']))
        
        for asignatura in data.get('evaluaciones_por_asignatura', []):
            story.append(Paragraph(
                f"<b>{asignatura['nombre']}</b>",
                self.styles['CustomNormal']
            ))
            story.append(Paragraph(
                f"Nota del trimestre: {asignatura['nota_trimestral']:.2f}",
                self.styles['CustomNormal']
            ))
            
            comentario_asignatura = comentarios.get('comentarios_asignaturas', {}).get(asignatura['nombre'], '')
            if comentario_asignatura:
                story.append(Paragraph(
                    comentario_asignatura,
                    self.styles['Comment']
                ))
            else:
                story.append(Paragraph(
                    "Comentario pendiente.",
                    self.styles['Comment']
                ))
            story.append(Spacer(1, 0.3*cm))
        
        # Observaciones sobre Asistencia
        if data.get('total_horas_ausencia', 0) > 10 and comentarios.get('comentario_asistencia'):
            story.append(Paragraph("🔹 Observaciones sobre Asistencia", self.styles['SectionTitle']))
            story.append(Paragraph(
                comentarios['comentario_asistencia'],
                self.styles['Comment']
            ))
            story.append(Spacer(1, 0.5*cm))
        
        # Footer
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph(
            f"Informe generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
            self.styles['CustomNormal']
        ))
        story.append(Paragraph(
            "EvalAI - Informes Inteligentes 2.0",
            self.styles['CustomNormal']
        ))
        
        # Generar PDF
        doc.build(story)
        buffer.seek(0)
        return buffer


class ExcelExportService:
    """
    Genera archivos Excel con datos crudos para análisis.
    """
    
    def generar_excel_grupo(self, data, trimestre, grupo_nombre):
        """
        Genera un Excel con los datos del grupo.
        
        Args:
            data: Diccionario con los datos del informe de grupo
            trimestre: T1, T2 o T3
            grupo_nombre: Nombre del grupo
            
        Returns:
            BytesIO con el Excel generado
        """
        
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        # Encabezados
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws_resumen['A1'] = f"Informe de Grupo: {grupo_nombre}"
        ws_resumen['A1'].font = Font(bold=True, size=14)
        ws_resumen['A2'] = f"Trimestre: {trimestre}"
        ws_resumen['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Datos generales
        ws_resumen['A5'] = "Indicador"
        ws_resumen['B5'] = "Valor"
        ws_resumen['A5'].fill = header_fill
        ws_resumen['B5'].fill = header_fill
        ws_resumen['A5'].font = header_font
        ws_resumen['B5'].font = header_font
        
        row = 6
        indicadores = [
            ('Total de Estudiantes', data.get('total_estudiantes', 0)),
            ('Media Global', f"{data.get('media_global', 0):.2f}"),
            ('Tasa de Aprobados', f"{data.get('tasa_aprobados', 0):.1f}%"),
            ('Asistencia Media', f"{data.get('asistencia_media', 0):.1f}%"),
            ('Total Horas de Falta', f"{data.get('total_horas_falta', 0):.0f}h"),
        ]
        
        for indicador, valor in indicadores:
            ws_resumen[f'A{row}'] = indicador
            ws_resumen[f'B{row}'] = valor
            row += 1
        
        # Hoja 2: Distribución de Notas
        ws_dist = wb.create_sheet("Distribución Notas")
        ws_dist['A1'] = "Categoría"
        ws_dist['B1'] = "Cantidad"
        ws_dist['C1'] = "Porcentaje"
        
        for col in ['A1', 'B1', 'C1']:
            ws_dist[col].fill = header_fill
            ws_dist[col].font = header_font
        
        row = 2
        for dist in data.get('distribucion_notas', []):
            ws_dist[f'A{row}'] = dist['categoria']
            ws_dist[f'B{row}'] = dist['cantidad']
            ws_dist[f'C{row}'] = f"{dist['porcentaje']}%"
            row += 1
        
        # Hoja 3: Medias por Asignatura
        ws_asig = wb.create_sheet("Medias Asignaturas")
        ws_asig['A1'] = "Asignatura"
        ws_asig['B1'] = "Media"
        ws_asig['C1'] = "Tendencia"
        
        for col in ['A1', 'B1', 'C1']:
            ws_asig[col].fill = header_fill
            ws_asig[col].font = header_font
        
        row = 2
        for asig in data.get('medias_por_asignatura', []):
            ws_asig[f'A{row}'] = asig['nombre']
            ws_asig[f'B{row}'] = round(asig['media'], 2)
            ws_asig[f'C{row}'] = round(asig['tendencia'], 2)
            row += 1
        
        # Hoja 4: Ranking Absentismo
        if data.get('ranking_absentismo'):
            ws_absent = wb.create_sheet("Ranking Absentismo")
            ws_absent['A1'] = "Posición"
            ws_absent['B1'] = "Estudiante"
            ws_absent['C1'] = "Horas de Ausencia"
            
            for col in ['A1', 'B1', 'C1']:
                ws_absent[col].fill = header_fill
                ws_absent[col].font = header_font
            
            row = 2
            for idx, est in enumerate(data['ranking_absentismo'], 1):
                ws_absent[f'A{row}'] = idx
                ws_absent[f'B{row}'] = est['nombre']
                ws_absent[f'C{row}'] = f"{est['horas_falta']}h"
                row += 1
        
        # Ajustar anchos de columna
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def generar_excel_individual(self, data, trimestre, estudiante_nombre):
        """
        Genera un Excel con los datos del estudiante.
        
        Args:
            data: Diccionario con los datos del estudiante
            trimestre: T1, T2 o T3
            estudiante_nombre: Nombre del estudiante
            
        Returns:
            BytesIO con el Excel generado
        """
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Estudiante"
        
        # Encabezados
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws['A1'] = f"Informe Individual: {estudiante_nombre}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Trimestre: {trimestre}"
        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Datos básicos
        ws['A5'] = "Dato"
        ws['B5'] = "Valor"
        ws['A5'].fill = header_fill
        ws['B5'].fill = header_fill
        ws['A5'].font = header_font
        ws['B5'].font = header_font
        
        ws['A6'] = "Grupo"
        ws['B6'] = data.get('grupo', '')
        ws['A7'] = "Total Horas de Ausencia"
        ws['B7'] = f"{data.get('total_horas_ausencia', 0)}h"
        
        # Evaluaciones por asignatura
        ws['A9'] = "Asignatura"
        ws['B9'] = "Nota Trimestral"
        ws['C9'] = "Tendencia"
        
        for col in ['A9', 'B9', 'C9']:
            ws[col].fill = header_fill
            ws[col].font = header_font
        
        row = 10
        for asig in data.get('evaluaciones_por_asignatura', []):
            ws[f'A{row}'] = asig['nombre']
            ws[f'B{row}'] = round(asig['nota_trimestral'], 2)
            ws[f'C{row}'] = round(asig['tendencia'], 2)
            row += 1
        
        # Autoevaluación
        if data.get('autoevaluacion'):
            ws[f'A{row + 2}'] = "Autoevaluación"
            ws[f'A{row + 2}'].font = Font(bold=True, size=12)
            ws[f'A{row + 3}'] = data['autoevaluacion']['texto']
        
        # Ajustar anchos
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# Instancias globales
pdf_export_service = PDFExportService()
excel_export_service = ExcelExportService()
